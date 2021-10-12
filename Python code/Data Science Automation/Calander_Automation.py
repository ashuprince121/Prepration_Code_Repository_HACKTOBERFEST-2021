from collections import Counter
from openpyxl.styles.borders import Border, Side
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def fetch_date(sh1, datee):
    # Fetch dates from the sheet
    for i in range(3, 34):
        j = 2
        datee.append(sh1.cell(i, j).value)
    return datee
    # Fetch month from the sheet

    # Fetch days from the sheet


def fetch_day(sh1, day):
    for i in range(3, 34):
        j = 3
        day.append(sh1.cell(i, j).value)
    return day

    # Fetch Course Code from the sheet without empty box


def fetch_CourseCode1(sh1, CC1):
    for i in range(3, 33):
        j = 4
        if (sh1.cell(i, j).value == None):

            continue
        else:
            CC1.append(sh1.cell(i, j).value)
    return CC1

    # Fetch Course code from sheet as it is


def fetch_CourseCode2(sh1, CC2):
    for i in range(3, 33):
        j = 4
        CC2.append(sh1.cell(i, j).value)
    return CC2

    # Fetch module from sheet


def fetch_module(sh1, module, day):
    m = 0
    for i in range(3, 33):
        j = 5
        if (day[m] == 'Sa' or day[m] == 'Su'):
            m = m + 1
            continue
        else:
            module.append(sh1.cell(i, j).value)
            m = m + 1
    return module

    # Fetch faculty1 from the sheet


def fetch_faculty1(sh1, faculty1, day):
    m = 0
    for i in range(3, 33):
        j = 6
        if (day[m] == 'Sa' or day[m] == 'Su'):
            m = m + 1
            continue
        elif (sh1.cell(i, j).value == None):
            faculty1.append(" ")
            m = m + 1
            continue
        else:
            faculty1.append(sh1.cell(i, j).value)
            m = m + 1
    return faculty1

    # Fetch Faculty 2 from the sheet


def fetch_faculty2(sh1, faculty2, day):
    m = 0
    for i in range(3, 33):
        j = 7
        if (day[m] == 'Sa' or day[m] == 'Su'):
            m = m + 1
            continue
        elif (sh1.cell(i, j).value == None):
            faculty2.append(" ")
            m = m + 1
            continue
        else:
            faculty2.append(sh1.cell(i, j).value)
            m = m + 1
    return faculty2


# Fetch Faculty 3 from the sheet
def fetch_faculty3(sh1, faculty3, day):
    m = 0
    for i in range(3, 33):
        j = 8
        if (day[m] == 'Sa' or day[m] == 'Su'):
            m = m + 1
            continue
        elif (sh1.cell(i, j).value == None):
            faculty3.append(" ")
            m = m + 1
            continue
        else:
            faculty3.append(sh1.cell(i, j).value)
            m = m + 1
    return faculty3


# Fetch Session slot from sheet
def fetch_SessionSlot(sh1, SS):
    for i in range(3, 34):
        j = 9
        SS.append(sh1.cell(i, j).value)
    return SS


# Creating new excel


# create sample table in new sheet
def FillLayout(sheet, month, thin_border, datee):
    sheet.cell(3, 1).value = 'Course Code'
    sheet.cell(3, 2).value = 'New Course Title'
    sheet.cell(3, 3).value = 'Faculty 1'
    sheet.cell(3, 4).value = 'Faculty 2'
    sheet.cell(3, 5).value = 'Faculty 3'
    # adjust cell width
    sheet.column_dimensions['A'].width = 14.29
    sheet.column_dimensions['B'].width = 72.29
    sheet.column_dimensions['C'].width = 15.71
    sheet.column_dimensions['D'].width = 15.71
    sheet.column_dimensions['E'].width = 15.71

    mon = ["", "Course Code", "New Course Title", "Faculty 1", "Faculty 2", "Faculty 3"]

    # merge cells and input month name in it with color and text arrangement
    sheet.merge_cells('F1:BO1')
    cell = sheet.cell(row=1, column=6)
    sheet['F1'].fill = PatternFill("solid", fgColor="FF7F50")
    cell.value = str(month)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    col_range = 5
    # colour the cells for date in new sheet
    for col in range(1, col_range + 1):
        cell_header = sheet.cell(3, col)
        cell_header.fill = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type="solid")
        sheet.cell(row=3, column=col).border = thin_border

    # enter the column name in new sheet
    for i in range(1, 6):
        r = 3
        cell = sheet.cell(row=r, column=i)
        cell.value = mon[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')

    newrow = len(datee) * 2
    # create border for coloured cells
    for col in range(6, newrow + 6, 2):
        cell_header = sheet.cell(3, col)
        cell_header.fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type="solid")
        sheet.cell(row=3, column=col).border = thin_border

    for col in range(7, newrow + 6, 2):
        cell_header = sheet.cell(3, col)
        cell_header.fill = PatternFill(start_color='6495ED', end_color='6495ED', fill_type="solid")
        sheet.cell(row=3, column=col).border = thin_border

    # writing session slot in new sheet


def SessionSlot(sheet, newrow):
    for i in range(6, newrow + 6, 2):
        r = 3
        cell = sheet.cell(row=r, column=i)
        cell.value = 'M'
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i in range(7, newrow + 6, 2):
        r = 3
        cell = sheet.cell(row=r, column=i)
        cell.value = 'A'
        cell.alignment = Alignment(horizontal='center', vertical='center')


# enter date in new sheet
def align(sheet, newrow):
    j = 1
    for i in range(6, newrow + 6, 2):
        r = 2
        for k in range(2):
            cell = sheet.cell(row=r, column=i)
            cell.value = j
            i = i + 1
            cell.alignment = Alignment(horizontal='center', vertical='center')
        j = j + 1


## Writing data

# course code
# CC =set(CC)
# remove repeated values from Course code list


# Enter course course in new sheet
def CourseCode(sheet, CC):
    r = 4
    for i in range(len(CC)):
        c = 1
        cell = sheet.cell(row=r, column=c)
        cell.value = CC[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        r = r + 1


# module =set(module)
# remove repeated values from module list


# Enter modules in new sheet
def Module(sheet, module):
    r = 4
    for i in range(len(module)):
        c = 2
        cell = sheet.cell(row=r, column=c)
        cell.value = module[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        r = r + 1


# enter Faculty 1 in new sheet

def Faculty1(sheet, faculty1):
    r = 4
    for i in range(len(faculty1)):
        c = 3
        cell = sheet.cell(row=r, column=c)
        cell.value = faculty1[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        r = r + 1


# enter faculty 2 in new sheet
def Faculty2(sheet, faculty2):
    r = 4
    for i in range(len(faculty2)):
        c = 4
        cell = sheet.cell(row=r, column=c)
        cell.value = faculty2[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        r = r + 1


# Enter faculty 3 in new sheet
def Faculty3(sheet, faculty3):
    r = 4
    for i in range(len(faculty3)):
        c = 5
        cell = sheet.cell(row=r, column=c)
        cell.value = faculty3[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        r = r + 1

    # Enter course code in new sheet


def ColourBlack(sheet, day, thin_border, CC):
    y = 6
    p = 0
    q = 4
    f = 0
    for i in range(1, len(day) + 1):
        x = 4

        if (day[p] == "Sa" or day[p] == "Su"):
            for d in range(2):
                for m in range(x, len(CC) + q):
                    cell_header = sheet.cell(x, y)
                    cell_header.fill = PatternFill(start_color='7C7C7B', end_color='7C7C7B', fill_type="solid")
                    sheet.cell(row=x, column=y).border = thin_border
                    x = x + 1
                y = y + 1
                x = 4

            p = p + 1

        else:
            y = y + 2
            p = p + 1


def ColourMA(sheet, SS, shift, thin_border):
    j = 0
    r = 4
    c = 6
    i = 0

    shift1 = []

    for h in range(len(shift)):
        n = shift[j]

        n = n + 2
        for m in range(n):
            if (i > (len(SS) - 1)):
                break
            if (SS[i] == None):
                n = n + 1
                i = i + 1
                c = c + 2
                continue
            elif (SS[i] == 'M&A'):
                for z in range(2):
                    cell = sheet.cell(row=r, column=c)
                    sheet.cell(row=r, column=c).border = thin_border
                    cell.fill = PatternFill(start_color='9FE2BF', end_color='9FE2BF', fill_type="solid")
                    c = c + 1
                i = i + 1
            elif (SS[i] == 'M'):
                cell = sheet.cell(row=r, column=c)
                cell.fill = PatternFill(start_color='9FE2BF', end_color='9FE2BF', fill_type="solid")
                sheet.cell(row=r, column=c).border = thin_border
                c = c + 2
                i = i + 1
            elif (SS[i] == 'A'):
                c = c + 1
                i = i + 1
                cell = sheet.cell(row=r, column=c)
                cell.fill = PatternFill(start_color='9FE2BF', end_color='9FE2BF', fill_type="solid")
                sheet.cell(row=r, column=c).border = thin_border

        r = r + 1
        j = j + 1


def mastercalander():
    shift = []
    datee = []
    day = []
    CC1 = []
    month = []
    module = []
    CC2 = []
    faculty1 = []
    faculty2 = []
    faculty3 = []
    SS = []
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # print("\n")
    # print("Enter the path to file in format - drive:\\...\\...\\filename.xlsx")
    str1="input_calendar.xlsx"

    wb = openpyxl.load_workbook(str1)
    sheets = wb.sheetnames
    sh1 = wb["Genesis"]
    row = sh1.max_row
    column = sh1.max_column
    wb = Workbook()
    sheet = wb.active

    datee = fetch_date(sh1, datee)

    month.append(sh1.cell(3, 1).value)

    day = fetch_day(sh1, day)

    CC1 = fetch_CourseCode1(sh1, CC1)

    CC2 = fetch_CourseCode2(sh1, CC2)

    module = fetch_module(sh1, module, day)

    faculty1.append(sh1.cell(4, 6).value)
    faculty1 = fetch_faculty1(sh1, faculty1, day)

    faculty2 = fetch_faculty2(sh1, faculty2, day)
    faculty2.append(sh1.cell(3, 7).value)

    faculty3 = fetch_faculty3(sh1, faculty3, day)
    faculty3.append(sh1.cell(3, 8).value)

    SS = fetch_SessionSlot(sh1, SS)

    FillLayout(sheet, month, thin_border, datee)

    newrow = len(datee) * 2
    SessionSlot(sheet, newrow)

    align(sheet, newrow)

    CC = list(dict.fromkeys(CC1))

    CourseCode(sheet, CC)

    module = list(dict.fromkeys(module))
    Module(sheet, module)

    faculty1 = list(dict.fromkeys(faculty1))
    Faculty1(sheet, faculty1)

    faculty2 = list(dict.fromkeys(faculty2))
    Faculty2(sheet, faculty2)

    faculty3 = list(dict.fromkeys(faculty3))
    Faculty3(sheet, faculty3)

    ColourBlack(sheet, day, thin_border, CC)

    # Enter colours in morning and afternnon slot in new sheet
    occ = Counter(CC1)
    for i in range(len(CC)):
        shift.append(occ[CC[i]])

    ColourMA(sheet, SS, shift, thin_border)

    ColourBlack(sheet, day, thin_border, CC)

    print("\nOutput Master Calander Excel File is Ready Now :)\n")
    wb.save("master.xlsx")



